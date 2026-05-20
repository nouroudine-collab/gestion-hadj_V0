import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook


DB_PATH = "data/app.db"


def _normalize(value: str) -> str:
    return (
        str(value or "")
        .strip()
        .lower()
        .replace("é", "e")
        .replace("è", "e")
        .replace("ê", "e")
        .replace("à", "a")
        .replace("ù", "u")
        .replace("ï", "i")
        .replace(" ", "")
        .replace("_", "")
    )


def _to_int(value, default=0):
    try:
        return int(float(str(value).replace(" ", "").replace(",", ".")))
    except Exception:
        return default


def import_excel_to_db(excel_path: str, db_path: str = DB_PATH):
    wb = load_workbook(excel_path, data_only=True)
    if "LISTE INSCRIPTION" in wb.sheetnames:
        ws = wb["LISTE INSCRIPTION"]
    else:
        ws = wb[wb.sheetnames[0]]

    headers = [cell.value for cell in ws[1]]
    idx = {_normalize(h): i for i, h in enumerate(headers) if h}

    def col(*candidates):
        for candidate in candidates:
            key = _normalize(candidate)
            if key in idx:
                return idx[key]
        return None

    i_nom = col("Nom et Prénom", "NOM PRENOM", "Nom Prenom")
    i_sexe = col("Sexe")
    i_date_naiss = col("Date de naissance", "DATE NAISS")
    i_lieu_naiss = col("Lieu de naissance", "LIEU NAISS")
    i_passport = col("N° Passeport / CNI-B", "PASSEPORT")
    i_deliv = col("Date de délivrance passeport", "DATE DELIV")
    i_auto = col("Versement AUTO (prévu)", "AUTO", "MONTANT PREVU")
    i_contact = col("Personne à prévenir", "PERSONNE A PREVENIR")
    i_tel1 = col("Téléphone 1", "Tel 1")
    i_tel2 = col("Téléphone 2", "Tel 2")

    with sqlite3.connect(db_path) as conn:
        cursor = conn.cursor()
        imported = 0
        skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if i_nom is None:
                break
            full_name = str(row[i_nom] or "").strip()
            if not full_name:
                skipped += 1
                continue

            parts = full_name.split()
            lname = parts[0].upper()
            fname = " ".join(parts[1:]).title() if len(parts) > 1 else "-"
            sex = str(row[i_sexe] or "M").strip().upper() if i_sexe is not None else "M"
            if sex not in ("M", "F"):
                sex = "M"
            birth_date = str(row[i_date_naiss] or "")[:10] if i_date_naiss is not None else ""
            birth_place = str(row[i_lieu_naiss] or "") if i_lieu_naiss is not None else ""
            passport = str(row[i_passport] or "") if i_passport is not None else ""
            deliv_date = str(row[i_deliv] or "")[:10] if i_deliv is not None else ""
            total_cost = _to_int(row[i_auto], default=0) if i_auto is not None else 0
            contact = str(row[i_contact] or "-") if i_contact is not None else "-"
            tel1 = str(row[i_tel1] or "").strip() if i_tel1 is not None else ""
            tel2 = str(row[i_tel2] or "").strip() if i_tel2 is not None else ""

            cursor.execute(
                "SELECT id FROM pilgrims WHERE UPPER(lname)=UPPER(?) AND UPPER(fname)=UPPER(?) AND birth_date=?",
                (lname, fname, birth_date),
            )
            if cursor.fetchone():
                skipped += 1
                continue

            cursor.execute("INSERT INTO persons_to_prevent(fullname) VALUES (?)", (contact,))
            person_id = cursor.lastrowid
            cursor.execute(
                """
                INSERT INTO pilgrims(lname, fname, sex, birth_date, birth_place, passport, deliv_date, total_cost, person_to_prevent_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (lname, fname, sex, birth_date, birth_place, passport, deliv_date, total_cost, person_id),
            )
            pilgrim_id = cursor.lastrowid
            if tel1:
                cursor.execute("INSERT INTO numbers(number, pilgrim_id) VALUES (?, ?)", (tel1, pilgrim_id))
            if tel2:
                cursor.execute("INSERT INTO numbers(number, pilgrim_id) VALUES (?, ?)", (tel2, pilgrim_id))
            imported += 1

        conn.commit()
    return imported, skipped


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python utils/import_hadj_excel.py <chemin_fichier_excel>")
        raise SystemExit(1)

    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"Fichier introuvable: {xlsx_path}")
        raise SystemExit(1)

    imported_count, skipped_count = import_excel_to_db(str(xlsx_path))
    print(f"Import termine. Inseres: {imported_count} | Ignores: {skipped_count}")
